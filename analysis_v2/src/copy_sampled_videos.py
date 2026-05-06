#!/usr/bin/env python3
"""
Script to copy video files based on sampled verification data.

This script:
1. Takes two paths as input: source videos folder and destination folder
2. Reads sampled_verification.json to get the list of files to copy
3. For each entry, extracts folder and subfolder names
4. Looks for video files in the source directory structure
5. Copies found files to destination, logs missing files
"""

import os
import json
import shutil
import argparse
import unicodedata
import re
import ffmpeg
from collections import defaultdict


def sanitize_name(name, replace_char='_'):
    """
    Sanitize a name by replacing spaces and hyphens with underscores.
    
    Args:
        name: The name to sanitize
        replace_char: Character to use as replacement for spaces and hyphens
        
    Returns:
        A sanitized version of the name
    """
    # Normalize Unicode characters (e.g., convert 'é' to 'e')
    name = unicodedata.normalize('NFKD', name)
    
    sanitized = name.replace(' ', replace_char).replace('-', replace_char).replace('._', replace_char)
    sanitized = re.sub(f'{replace_char}+', replace_char, sanitized)
    sanitized = sanitized.strip(replace_char)
    return sanitized

def canonicalize_for_match(name):
    """
    Canonicalize names for path/file matching.

    This intentionally removes separator differences so that names with
    '-' and '_' match each other.
    """
    normalized = unicodedata.normalize('NFKD', str(name or ""))
    normalized = normalized.lower()
    # Treat separators the same.
    normalized = re.sub(r'[\s._-]+', '', normalized)
    # Drop remaining non-alphanumeric characters for robust matching.
    normalized = re.sub(r'[^a-z0-9]', '', normalized)
    return normalized


def parse_sampled_verification_json(json_path):
    """
    Parse the sampled_verification.json file and extract file paths.
    
    Args:
        json_path (str): Path to the sampled_verification.json file
        
    Returns:
        dict: Dictionary with folder structure and file paths
    """
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
        return data
    except FileNotFoundError:
        print(f"Error: Could not find {json_path}")
        return {}
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in {json_path}")
        return {}

def parse_sampled_verification_xlsx(xlsx_path, sheet_name=None):
    """
    Parse sampled verification data from an Excel workbook.

    Expected columns:
      - conf
      - session
      - video
      - chunk_file

    Returns a dictionary compatible with the JSON parser output:
      {
        "conf/session": {
          "video/chunk_file.json": {}
        }
      }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: Missing dependency 'openpyxl'. Install it with: pip install openpyxl")
        return {}

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"Error: Could not find {xlsx_path}")
        return {}
    except Exception as e:
        print(f"Error: Could not read Excel file {xlsx_path}: {e}")
        return {}

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            return {}
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        print(f"Error: Excel file {xlsx_path} is empty")
        return {}

    header_to_idx = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_to_idx[str(header).strip().lower()] = idx

    required_columns = ['conf', 'session', 'video', 'chunk_file']
    missing_columns = [col for col in required_columns if col not in header_to_idx]
    if missing_columns:
        print(f"Error: Missing required columns in Excel file: {missing_columns}")
        return {}

    data = defaultdict(dict)
    seen = set()

    for row in rows:
        conf = row[header_to_idx['conf']]
        session = row[header_to_idx['session']]
        video = row[header_to_idx['video']]
        chunk_file = row[header_to_idx['chunk_file']]

        if not conf or not session or not video or not chunk_file:
            continue

        conf = str(conf).strip()
        session = str(session).strip()
        video = os.path.basename(str(video).strip())
        chunk_file = os.path.basename(str(chunk_file).strip())

        # Deduplicate repeated transcript rows that point to the same chunk.
        dedupe_key = (conf, session, video, chunk_file)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        key = f"{conf}/{session}"
        value_key = f"{video}/{chunk_file}"
        data[key][value_key] = {}

    return dict(data)


def parse_sampled_verification(input_path, sheet_name=None):
    """
    Parse sampled verification data from JSON or XLSX.
    """
    extension = os.path.splitext(input_path)[1].lower()
    if extension == '.json':
        return parse_sampled_verification_json(input_path)
    if extension in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return parse_sampled_verification_xlsx(input_path, sheet_name=sheet_name)

    print(f"Error: Unsupported input file extension '{extension}'. Use .json or .xlsx")
    return {}


def find_matching_directory(parent_dir, target_name):
    """
    Find a directory under parent_dir whose name matches target_name, allowing
    separator differences like '_' vs '-'.
    """
    if not os.path.isdir(parent_dir):
        return None

    exact = os.path.join(parent_dir, target_name)
    if os.path.isdir(exact):
        return exact

    target_canonical = canonicalize_for_match(target_name)
    for entry in os.listdir(parent_dir):
        entry_path = os.path.join(parent_dir, entry)
        if os.path.isdir(entry_path) and canonicalize_for_match(entry) == target_canonical:
            return entry_path
    return None


def find_matching_video_file(folder_path, file_name, allowed_extensions):
    """
    Find a video file in folder_path that matches file_name with allowed
    separator differences.
    """
    target_canonical = canonicalize_for_match(file_name)

    # Exact-name check first.
    for ext in allowed_extensions:
        candidate = os.path.join(folder_path, file_name + ext)
        if os.path.exists(candidate):
            return candidate, ext

    for entry in os.listdir(folder_path):
        entry_path = os.path.join(folder_path, entry)
        if not os.path.isfile(entry_path):
            continue
        stem, ext = os.path.splitext(entry)
        if ext.lower() not in allowed_extensions:
            continue
        if canonicalize_for_match(stem) == target_canonical:
            return entry_path, ext.lower()
    return None, None


def find_matching_split_folder(folder_path, file_name):
    """
    Find split folder for a file_name, accepting both split_<name> and
    split-<name> (and other separator variants).
    """
    direct_candidates = [
        os.path.join(folder_path, f"split_{file_name}"),
        os.path.join(folder_path, f"split-{file_name}")
    ]
    for candidate in direct_candidates:
        if os.path.isdir(candidate):
            return candidate

    target_canonical = canonicalize_for_match(file_name)
    for entry in os.listdir(folder_path):
        entry_path = os.path.join(folder_path, entry)
        if not os.path.isdir(entry_path):
            continue

        if not entry.lower().startswith("split"):
            continue

        remainder = entry[5:].lstrip("_- ")
        if canonicalize_for_match(remainder) == target_canonical:
            return entry_path
    return None


def find_matching_chunk_file(split_folder_path, chunk_file_name):
    """
    Find chunk file under split folder, ignoring '_' vs '-' differences.
    """
    exact = os.path.join(split_folder_path, chunk_file_name + '.mp4')
    if os.path.exists(exact):
        return exact

    target_canonical = canonicalize_for_match(chunk_file_name)
    for entry in os.listdir(split_folder_path):
        entry_path = os.path.join(split_folder_path, entry)
        if not os.path.isfile(entry_path):
            continue
        stem, ext = os.path.splitext(entry)
        if ext.lower() != '.mp4':
            continue
        if canonicalize_for_match(stem) == target_canonical:
            return entry_path
    return None


def extract_folder_info(key):
    """
    Extract folder and subfolder information from a key like "2021MND/output_2021_04_22_MND_S6".
    
    Args:
        key (str): Key from sampled_verification.json
        
    Returns:
        tuple: (folder_name, subfolder_name)
    """
    parts = key.split('/')
    if len(parts) != 2:
        return None, None
    
    folder_name = sanitize_name(parts[0])
      # e.g., "2021MND"
    
    # Extract subfolder name as what follows after "output_"
    if parts[1].startswith('output_'):
        subfolder_name = parts[1][7:]  # Remove "output_" prefix (7 characters)
    else:
        # If it doesn't start with "output_", use the whole part
        subfolder_name = parts[1]
    
    return folder_name, subfolder_name

# convert mkv to mp4 using ffmpeg
def convert_mkv_to_mp4(input_path):
    """
    Converts an MKV video file to MP4 using ffmpeg.
    Re-encodes audio to AAC to ensure compatibility.
    """
    import subprocess
    import os

    if not input_path.lower().endswith(".mkv"):
        raise ValueError("Input file must be an MKV file.")
    
    output_path = os.path.splitext(input_path)[0] + ".mp4"

    command = [
        "ffmpeg",
        "-i", input_path,
        "-c:v", "copy",        # Copy video stream
        "-c:a", "aac",         # Re-encode audio to AAC
        "-b:a", "192k",        # Set audio bitrate (optional)
        output_path
    ]

    try:
        subprocess.run(command, check=True)
        print(f"Conversion successful: {output_path}")
        return output_path
    except subprocess.CalledProcessError as e:
        print(f"Conversion failed: {e}")
        return None

# Split a video into chunks of specified length
def split_video(video_full_path, duration, chunk_length=10*60):
    
    # Calculate the number of chunks
    num_chunks = int(duration // chunk_length) + 1
    
    # Get the file name and directory
    file_name, _ = os.path.splitext(os.path.basename(video_full_path))
    directory = os.path.dirname(video_full_path)
    
    # Create a directory to store the split videos
    split_dir = os.path.join(directory, f"split_{file_name}")
    os.makedirs(split_dir, exist_ok=True)
    
    # Split the video into chunks
    chunk_paths = []
    for i in range(num_chunks):
        start_time = i * chunk_length
        output_file_name = os.path.join(split_dir, f"{file_name}_chunk{i+1}.mp4")
        # Add -n flag to skip existing files
        ffmpeg.input(video_full_path, ss=start_time, t=chunk_length).output(output_file_name, n=None).run()
        chunk_paths.append(output_file_name)
        print(f"Created chunk: {output_file_name}")
    
    return chunk_paths

def extract_file_info(value):
    """
    Extract file information from a value like "Breakout_Room_4_Part_2_2021_04_22_13_14_53/Breakout_Room_4_Part_2_2021_04_22_13_14_53_chunk6.json".
    
    Args:
        value (str): Value from sampled_verification.json
        
    Returns:
        tuple: (file_name, chunk_file_name)
    """
    parts = value.split('/')
    if len(parts) != 2:
        return None, None
    
    file_name = sanitize_name(parts[0])  # e.g., "Breakout_Room_4_Part_2_2021_04_22_13_14_53"
    
    # Remove .json extension from chunk file name
    chunk_file_name = sanitize_name(parts[1])  # e.g., "Breakout_Room_4_Part_2_2021_04_22_13_14_53_chunk6.json"
    if chunk_file_name.endswith('.json'):
        chunk_file_name = chunk_file_name[:-5]  # Remove ".json" (5 characters)
    
    return file_name, chunk_file_name


def find_video_file(source_dir, folder_name, subfolder_name, file_name, chunk_file_name, verbose=False):
    """
    Find the video file in the source directory structure.
    
    Args:
        source_dir (str): Source directory containing all videos
        folder_name (str): Main folder name (e.g., "2021MND")
        subfolder_name (str): Subfolder name (e.g., "output_2021_04_22_MND_S6")
        file_name (str): File name without extension
        chunk_file_name (str): Chunk file name
        verbose (bool): Whether to print detailed information
        
    Returns:
        str: Full path to the video file if found, None otherwise
    """
    # Construct the path to look for the video file, allowing separator
    # differences in folder names.
    folder_root = find_matching_directory(source_dir, folder_name)
    if not folder_root:
        if verbose:
            print(f"      Folder does not exist: {os.path.join(source_dir, folder_name)}")
        return None

    folder_path = find_matching_directory(folder_root, subfolder_name)
    
    if verbose:
        print(f"      Looking in folder: {folder_path}")
    
    if not folder_path or not os.path.exists(folder_path):
        if verbose:
            print(f"      Subfolder does not exist: {os.path.join(folder_root, subfolder_name)}")
        return None
    
    # Look for the video file (with various extensions)
    video_extensions = ['.mp4', '.mkv', '.avi', '.mov', '.wmv']
    video_file_path, video_extension = find_matching_video_file(folder_path, file_name, video_extensions)
    if video_file_path and verbose:
        print(f"      Found video file: {video_file_path}")
    
    if not video_file_path:
        if verbose:
            print(f"      No video file found for: {file_name}")
        return None
    
    # Look for the split folder (allow split_<name> vs split-<name>)
    split_folder_path = find_matching_split_folder(folder_path, file_name)
    
    if verbose:
        print(f"      Looking for split folder: {split_folder_path}")
    
    if not split_folder_path or not os.path.exists(split_folder_path):
        if verbose:
            print(f"      Split folder does not exist for: {file_name}")
            print(f"      Will create split folder and process video")
        
        # Check if the video file is MKV and needs conversion
        if video_extension.lower() == '.mkv':
            if verbose:
                print(f"      Converting MKV to MP4: {video_file_path}")
            try:
                converted_path = convert_mkv_to_mp4(video_file_path)
                if converted_path and os.path.exists(converted_path):
                    video_file_path = converted_path
                    if verbose:
                        print(f"      Successfully converted to: {converted_path}")
                else:
                    if verbose:
                        print(f"      Failed to convert MKV file")
                    return None
            except Exception as e:
                if verbose:
                    print(f"      Error converting MKV: {e}")
                return None
        
        # Split the video into chunks
        if verbose:
            print(f"      Splitting video: {video_file_path}")
        try:
            # Get video duration using ffmpeg
            probe = ffmpeg.probe(video_file_path)
            duration = None
            for stream in probe.get('streams', []):
                if 'duration' in stream:
                    duration = float(stream['duration'])
                    break
            if duration is None and 'format' in probe and 'duration' in probe['format']:
                duration = float(probe['format']['duration'])
            if duration is None:
                raise ValueError("Could not determine video duration")
            
            # Split the video
            chunk_paths = split_video(video_file_path, duration)
            split_folder_path = os.path.dirname(chunk_paths[0]) if chunk_paths else None
            if verbose:
                print(f"      Successfully created {len(chunk_paths)} chunks")
        except Exception as e:
            if verbose:
                print(f"      Error splitting video: {e}")
            return None
    
    # Look for the chunk file in the split folder
    chunk_file_path = find_matching_chunk_file(split_folder_path, chunk_file_name)
    
    if verbose:
        print(f"      Looking for chunk file: {chunk_file_path}")
    
    if chunk_file_path and os.path.exists(chunk_file_path):
        if verbose:
            print(f"      Found chunk file: {chunk_file_path}")
        return chunk_file_path
    
    if verbose:
        print(f"      Chunk file not found: {chunk_file_path}")
    return None


def copy_video_files(source_dir, dest_dir, input_path, sheet_name=None, dry_run=False, verbose=False):
    """
    Copy video files based on sampled verification data.
    
    Args:
        source_dir (str): Source directory containing all videos
        dest_dir (str): Destination directory for copied files
        input_path (str): Path to sampled_verification input file (.json or .xlsx)
        sheet_name (str): Optional Excel sheet name
        dry_run (bool): Whether to simulate the operation without making changes
        verbose (bool): Whether to print detailed information
        
    Returns:
        tuple: (copied_count, missing_files)
    """
    # Create destination directory if it doesn't exist
    if not dry_run:
        os.makedirs(dest_dir, exist_ok=True)
    elif verbose:
        print(f"Would create destination directory: {dest_dir}")
    
    # Parse the JSON file
    data = parse_sampled_verification(input_path, sheet_name=sheet_name)
    if not data:
        return 0, []
    
    copied_count = 0
    missing_files = []
    
    print(f"Processing {len(data)} entries from {input_path}")
    
    for key, values in data.items():
        print(f"\nProcessing key: {key}")
        
        # Extract folder information
        folder_name, subfolder_name = extract_folder_info(key)
        if not folder_name or not subfolder_name:
            print(f"  Skipping invalid key format: {key}")
            continue
        
        if verbose:
            print(f"  Folder: {folder_name}, Subfolder: {subfolder_name}")
        
        # Process each value in the key
        if isinstance(values, dict):
            # If values is a dictionary, process each key-value pair
            for value_key, value_data in values.items():
                file_name, chunk_file_name = extract_file_info(value_key)
                if not file_name or not chunk_file_name:
                    print(f"    Skipping invalid value format: {value_key}")
                    continue
                
                if verbose:
                    print(f"    File: {file_name}, Chunk: {chunk_file_name}")
                
                # Find the video file
                video_file_path = find_video_file(source_dir, folder_name, subfolder_name, file_name, chunk_file_name, verbose)
                
                if video_file_path:
                    # Copy the file
                    dest_file_path = os.path.join(dest_dir, chunk_file_name + '.mp4')
                    
                    # Check if file already exists in destination
                    if os.path.exists(dest_file_path):
                        if verbose:
                            print(f"    Skipping: {chunk_file_name}.mp4 (already exists)")
                        copied_count += 1
                        continue
                    
                    if dry_run:
                        print(f"    Would copy: {video_file_path} -> {dest_file_path}")
                        copied_count += 1
                    else:
                        try:
                            shutil.copy2(video_file_path, dest_file_path)
                            print(f"    Copied: {chunk_file_name}.mp4")
                            copied_count += 1
                        except Exception as e:
                            print(f"    Error copying {chunk_file_name}.mp4: {e}")
                            missing_files.append(f"{key}/{value_key}")
                else:
                    print(f"    Not found: {chunk_file_name}")
                    missing_files.append(f"{key}/{value_key}")
        else:
            # If values is not a dictionary, skip
            print(f"  Skipping non-dictionary values for key: {key}")
    
    return copied_count, missing_files


def save_missing_files(missing_files, output_path, dry_run=False):
    """
    Save missing files to a JSON file.
    
    Args:
        missing_files (list): List of missing file paths
        output_path (str): Path to save the JSON file
        dry_run (bool): Whether to simulate the operation without making changes
    """
    if dry_run:
        print(f"Would save missing files to: {output_path}")
        print(f"Missing files: {missing_files}")
    else:
        with open(output_path, 'w') as f:
            json.dump(missing_files, f, indent=2)
        print(f"Missing files saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Copy video files based on sampled verification data')
    parser.add_argument('source_dir', help='Path to folder containing all videos')
    parser.add_argument('dest_dir', help='Path to folder where copied files should be stored')
    parser.add_argument('--input-file', default='sampling/sampled_verification.json',
                       help='Path to sampled verification file (.json or .xlsx)')
    # Backward-compatible alias for older invocations.
    parser.add_argument('--json-file', dest='input_file', help=argparse.SUPPRESS)
    parser.add_argument('--sheet-name', default=None,
                       help='Excel sheet name to read (default: active sheet)')
    parser.add_argument('--missing-output', default='missing_files.json',
                       help='Path to save missing files JSON (default: missing_files.json)')
    parser.add_argument('--dry-run', action='store_true', help='Print what would be done without making changes')
    parser.add_argument('--verbose', '-v', action='store_true', help='Print detailed information')
    
    args = parser.parse_args()
    
    # Validate source directory
    if not os.path.exists(args.source_dir):
        print(f"Error: Source directory does not exist: {args.source_dir}")
        return 1
    
    # Validate input file
    if not os.path.exists(args.input_file):
        print(f"Error: Input file does not exist: {args.input_file}")
        return 1
    
    print(f"Source directory: {args.source_dir}")
    print(f"Destination directory: {args.dest_dir}")
    print(f"Input file: {args.input_file}")
    if args.sheet_name:
        print(f"Excel sheet: {args.sheet_name}")
    if args.dry_run:
        print("DRY RUN - no changes will be made")
    if args.verbose:
        print("VERBOSE - detailed information will be shown")
    print("-" * 50)
    
    # Copy video files
    copied_count, missing_files = copy_video_files(
        args.source_dir,
        args.dest_dir,
        args.input_file,
        sheet_name=args.sheet_name,
        dry_run=args.dry_run,
        verbose=args.verbose
    )
    
    # Save missing files
    if missing_files:
        save_missing_files(missing_files, args.missing_output, args.dry_run)
    
    # Print summary
    print("\n" + "=" * 50)
    print(f"Summary:")
    if args.dry_run:
        print(f"  Files that would be copied: {copied_count}")
    else:
        print(f"  Files copied: {copied_count}")
    print(f"  Files missing: {len(missing_files)}")
    if not args.dry_run:
        print(f"  Missing files saved to: {args.missing_output}")
    
    return 0


if __name__ == "__main__":
    exit(main()) 
